import sys
import re
from datetime import datetime
from email.utils import parsedate_to_datetime
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
import os
import base64
from googleapiclient.discovery import build
import openpyxl
import subprocess

# Gmail API Scopes (adjust as needed)
SCOPES = [
    'https://www.googleapis.com/auth/gmail.modify',
    'https://www.googleapis.com/auth/gmail.labels',
    'https://www.googleapis.com/auth/gmail.readonly'
]


def parse_bool_field(value, rule_name, field_name):
    """
    For boolean fields (both conditions and actions), allow only: True, False, or empty.
    Empty is interpreted as True.
    Otherwise, throw a ValueError indicating which rule has an error.
    """
    if value is None or str(value).strip() == "":
        return True
    normalized = str(value).strip().lower()
    if normalized == "true":
        return True
    elif normalized == "false":
        return False
    else:
        raise ValueError(f"Error in rule '{rule_name}': Field '{field_name}' must be True, False, or empty. Got: {value}")


def get_gmail_service():
    print("Initializing Gmail service...")
    creds = None
    token_path = 'resources/token.json'
    credentials_path = 'resources/credentials.json'

    print(f"Checking for existing token at {token_path}")
    if os.path.exists(token_path):
        print("Found existing token, loading credentials...")
        creds = Credentials.from_authorized_user_file(token_path, SCOPES)
    else:
        print("No existing token found.")
    
    if not creds or not creds.valid:
        print("Credentials are invalid or expired.")
        if creds and creds.expired and creds.refresh_token:
            print("Attempting to refresh expired token...")
            creds.refresh(Request())
            print("Token refreshed successfully.")
        else:
            print("Need to generate new token...")
            flow = InstalledAppFlow.from_client_secrets_file(
                credentials_path,
                SCOPES
            )
            creds = flow.run_local_server(port=0)
            print("New token generated.")
        
        print(f"Saving token to {token_path}")
        with open(token_path, 'w') as token:
            token.write(creds.to_json())
    
    print("Building Gmail service...")
    service = build('gmail', 'v1', credentials=creds)
    print("Gmail service initialized successfully!")
    return service


def load_rules_from_excel(file_path, sheet_name=None):
    """Load rules from an Excel file from a specified sheet.
    Sheet structure:
      1: Name (irrelevant)
      2: Description (irrelevant)
      3: Category (irrelevant)
      4: Rule Type
      5: Condition Type 1
      6: Condition Value 1
      7: Condition Type 2
      8: Condition Value 2
      9: Condition Type 3
      10: Condition Value 3
      11: Action
      12: Action Value
    """
    print("Loading rules from Excel...")
    rules = []
    workbook = openpyxl.load_workbook(file_path)
    
    if sheet_name and sheet_name in workbook.sheetnames:
        print(f"Loading rules from sheet: {sheet_name}")
        sheet = workbook[sheet_name]
    else:
        print("Sheet not specified or not found; using active sheet.")
        sheet = workbook.active

    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):  # Skip header
        rule_type = row[3]
        if not rule_type or (isinstance(rule_type, str) and (rule_type.lower() == 'none' or rule_type.lower() == 'paused')):
            continue

        rule = {
            'name': str(row[0]) if row[0] else f"Rule_{row_idx}",
            'rule_type': str(rule_type).lower() if rule_type else None,
            'conditions': [],
            'action': str(row[10]).lower() if row[10] and str(row[10]).lower() != 'none' else None,
            'action_value': str(row[11]) if row[11] and str(row[11]).lower() != 'none' else None
        }

        condition_pairs = [
            (row[4], row[5]),  # Condition 1
            (row[6], row[7]),  # Condition 2
            (row[8], row[9])   # Condition 3
        ]

        for cond_type, cond_value in condition_pairs:
            if cond_type and cond_value and str(cond_type).lower() != 'none':
                rule['conditions'].append({
                    'type': str(cond_type).lower(),
                    'value': str(cond_value).lower()
                })
        
        # print(f"Loaded rule: {rule['name']}")
        # print(f"  Rule Type: {rule['rule_type']}")
        # print(f"  Conditions: {len(rule['conditions'])}")
        # print(f"  Action: {rule['action']} -> {rule['action_value']}")
        rules.append(rule)
    
    print("\n=== Rules Loaded from Excel ===")
    for idx, rule in enumerate(rules, start=1):
        print(f"Rule {idx}:")
        print(f"  Name: {rule['name']}")
        # print(f"  Rule Type: {rule['rule_type']}")
        # print(f"  Conditions:")
        # for c_idx, cond in enumerate(rule['conditions'], start=1):
        #     print(f"    Condition {c_idx}: type='{cond['type']}', value='{cond['value']}'")
        # print(f"  Action: {rule['action']}")
        # print(f"  Action Value: {rule['action_value']}")
    print("================================\n")

    return rules


def get_full_body_text(payload):
    """
    Recursively extract all text (text/plain or text/html) from the message payload,
    decode from base64, and return as a single string.
    """
    parts_text = []

    def walk(part):
        mime_type = part.get('mimeType', '')
        # If this part has sub-parts, walk them recursively
        if 'parts' in part:
            for p in part['parts']:
                walk(p)
        else:
            body = part.get('body', {})
            data = body.get('data')
            # Only decode if there's actual data and the MIME type is text
            if data and ('text/plain' in mime_type or 'text/html' in mime_type):
                decoded = base64.urlsafe_b64decode(data).decode('utf-8', errors='replace')
                parts_text.append(decoded)

    walk(payload)
    # Combine all text parts with a newline
    return "\n".join(parts_text)


def evaluate_conditions(conditions, subject, sender, body, email_date, labels, service, is_important=False, attachments=None, replies_count=0):
    subject_lower = subject.lower() if subject else ''
    sender_lower = sender.lower() if sender else ''
    body_lower = body.lower() if body else ''

    # (Debug prints removed here so that they don’t run for each rule)

    for idx, condition in enumerate(conditions, 1):
        cond_type = condition['type'].lower()
        cond_value = condition['value'].strip()

        if cond_type == 'none':
            continue

        if cond_type == 'sender':
            field_value = sender_lower
            if cond_value.startswith('[') and cond_value.endswith(']'):
                expression = cond_value[1:-1]
                result = evaluate_logical_expression(expression, field_value)
                if not result:
                    return False
            else:
                if cond_value not in field_value:
                    return False

        elif cond_type == 'subject':
            field_value = subject_lower
            if cond_value.startswith('[') and cond_value.endswith(']'):
                expression = cond_value[1:-1]
                result = evaluate_logical_expression(expression, field_value)
                if not result:
                    return False
            else:
                if cond_value not in field_value:
                    return False

        elif cond_type == 'body':
            field_value = body_lower
            if cond_value.startswith('[') and cond_value.endswith(']'):
                expression = cond_value[1:-1]
                result = evaluate_logical_expression(expression, field_value)
                if not result:
                    return False
            else:
                if cond_value not in field_value:
                    return False

        elif cond_type == 'date':
            m = re.match(r'^\[\s*(<|<=|>|>=|==)\s*([\d]{4}-[\d]{2}-[\d]{2})\s*\]$', cond_value)
            if m:
                op, cond_date_str = m.groups()
                cond_date = datetime.strptime(cond_date_str, '%Y-%m-%d')
                if op == '<' and not (email_date < cond_date):
                    return False
                elif op == '<=' and not (email_date <= cond_date):
                    return False
                elif op == '>' and not (email_date > cond_date):
                    return False
                elif op == '>=' and not (email_date >= cond_date):
                    return False
                elif op == '==' and not (email_date == cond_date):
                    return False
            else:
                if cond_value.lower().strip('[]') not in email_date.strftime('%Y-%m-%d').lower():
                    return False

        elif cond_type == 'important':
            try:
                cond_bool = parse_bool_field(cond_value, rule_name=condition.get('name','<unknown>'), field_name="important condition")
            except ValueError as e:
                print(e)
                return False
            if cond_bool != is_important:
                return False

        elif cond_type == 'tag':
            if cond_value in ["true", "false", ""]:
                try:
                    cond_bool = parse_bool_field(cond_value, rule_name=condition.get('name','<unknown>'), field_name="tag condition")
                except ValueError as e:
                    print(e)
                    return False
                print(f"Error: Boolean tag conditions are not supported (Rule '{condition.get('name','<unknown>')}')")
                return False
            else:
                if labels is None or service is None:
                    return False
                tag_label_id = get_or_create_label(service, cond_value)
                if tag_label_id not in labels:
                    return False

        elif cond_type == 'attachment':
            if cond_value in ["true", "false", ""]:
                try:
                    cond_bool = parse_bool_field(cond_value, rule_name=condition.get('name','<unknown>'), field_name="attachment condition")
                except ValueError as e:
                    print(e)
                    return False
                attachments_exist = bool(attachments)
                if cond_bool != attachments_exist:
                    return False
            else:
                attachments_str = ' '.join(attachments).lower() if attachments else ''
                if cond_value.lower() not in attachments_str:
                    return False

        elif cond_type == 'replies':
            m = re.match(r'^([<>])\s*(\d+)$', cond_value)
            if not m:
                print(f"Invalid format for replies condition: {cond_value}")
                return False
            op, num_str = m.groups()
            required_count = int(num_str)
            if op == '>' and not (replies_count > required_count):
                return False
            elif op == '<' and not (replies_count < required_count):
                return False
            elif op == '==' and not (replies_count == required_count):
                return False

        else:
            field_value = subject_lower
            if cond_value.startswith('[') and cond_value.endswith(']'):
                expression = cond_value[1:-1]
                result = evaluate_logical_expression(expression, field_value)
                if not result:
                    return False
            else:
                if cond_value not in field_value:
                    return False

    return True


def evaluate_logical_expression(expression, text):
    """
    Evaluate a logical expression (supports AND, OR, XOR, NOT, and parentheses)
    on the given text. This function assumes that the expression is provided without
    the enclosing square brackets (i.e. user wrote ["foo" and "bar"], we receive
    '"foo" and "bar"' here).
    """

    # Trim surrounding whitespace
    expression = expression.strip()

    expression = re.sub(r'[\r\n]+', ' ', expression)
    expression = re.sub(r'\s+', ' ', expression)

    def strip_outer_parens(expr):
        """
        If the entire expr is enclosed in matching parentheses, strip them
        and return the inside. Otherwise, return expr unchanged.
        """
        expr = expr.strip()
        if len(expr) < 2:
            return expr
        
        # Must start with '(' and end with ')'
        if not (expr.startswith('(') and expr.endswith(')')):
            return expr
        
        # Check if the first '(' matches the final ')'
        paren_count = 0
        for i, ch in enumerate(expr):
            if ch == '(':
                paren_count += 1
            elif ch == ')':
                paren_count -= 1
                # If paren_count hits 0 before the end, they're not wrapping the whole expr
                if paren_count == 0 and i < len(expr) - 1:
                    return expr
        # If we got here, the outer parentheses do wrap the entire expression
        return expr[1:-1].strip()

    def parse_expr(expr):
        expr = expr.strip()
        
        # 1) Strip outer parentheses if they truly enclose the entire subexpression
        old_expr = expr
        new_expr = strip_outer_parens(expr)
        while new_expr != expr:
            expr = new_expr
            new_expr = strip_outer_parens(expr)

        # 2) Handle leading NOT operator
        #    e.g. "not (foo or bar)" => not <subexpression>
        #    e.g. "not foo" => not <keyword>
        if expr.startswith('not '):
            return not parse_expr(expr[4:].strip())

        # 3) Scan for top-level operators (AND, OR, XOR) outside parentheses/quotes
        in_quotes = False
        paren_level = 0
        op_index = None
        op_found = None
        i = 0
        while i < len(expr):
            ch = expr[i]
            if ch == '"':
                in_quotes = not in_quotes
            elif not in_quotes:
                if ch == '(':
                    paren_level += 1
                elif ch == ')':
                    paren_level -= 1
                # Only check for operators at top level (paren_level == 0)
                if paren_level == 0:
                    # Check for multi-char operators with surrounding spaces
                    for op in [' and ', ' or ', ' xor ']:
                        if expr[i:].startswith(op):
                            op_index = i
                            op_found = op.strip()  # 'and', 'or', or 'xor'
                            break
                    if op_found:
                        break
            i += 1

        if op_found:
            # We found a top-level operator at op_index
            left_str = expr[:op_index]
            # The operator has length len(op_found) + 2 spaces = e.g. ' and ' => 5 chars
            op_len = len(op_found) + 2  # e.g. ' and ' is 5 chars
            right_str = expr[op_index + op_len:]

            left_val = parse_expr(left_str)
            right_val = parse_expr(right_str)

            if op_found == 'and':
                return left_val and right_val
            elif op_found == 'or':
                return left_val or right_val
            elif op_found == 'xor':
                return bool(left_val) != bool(right_val)

        # 4) No top-level operator found => treat as a simple keyword search
        #    Remove any leading/trailing quotes
        keyword = expr.strip().strip('"\'')
        return keyword in text

    # Kick off parsing
    return parse_expr(expression)


def get_or_create_label(service, label_name):
    """Get existing label ID or create a new one."""
    try:
        results = service.users().labels().list(userId='me').execute()
        labels = results.get('labels', [])
        for label in labels:
            if label['name'].lower() == label_name.lower():
                # print("label found")
                return label['id']
            
        # Create new label if not found.
        # print("creating new label")
        new_label = {
            'name': label_name,
            'labelListVisibility': 'labelShow',
            'messageListVisibility': 'show'
        }
        created = service.users().labels().create(userId='me', body=new_label).execute()
        return created['id']
    except Exception as e:
        print(f"Error handling labels: {e}")
        return None


def handle_rule_action(service, message_id, rule):
    """Execute the action specified in the rule."""
    action = rule['action']
    action_value = rule['action_value']
    rule_name = rule.get('name', 'Unknown Rule')

    if not action or action.lower() == 'none':
        print("No action specified")
        return

    action_lower = action.lower()

    try:
        if action_lower in ['mark_important', 'star', 'archive', 'delete']:
            action_bool = parse_bool_field(action_value, rule_name, f"action {action_lower}")
    except ValueError as e:
        print(e)
        return

    try:
        if action_lower == 'mark_important':
            if action_bool:
                service.users().messages().modify(
                    userId='me',
                    id=message_id,
                    body={
                        'addLabelIds': ['IMPORTANT'],
                        'removeLabelIds': []
                    }
                ).execute()
                print(f"Marked email as important (Rule: {rule_name})")
            else:
                service.users().messages().modify(
                    userId='me',
                    id=message_id,
                    body={
                        'addLabelIds': [],
                        'removeLabelIds': ['IMPORTANT']
                    }
                ).execute()
                print(f"Marked email as not important (Rule: {rule_name})")

        elif action_lower == 'star':
            if action_bool:
                service.users().messages().modify(
                    userId='me',
                    id=message_id,
                    body={
                        'addLabelIds': ['STARRED'],
                        'removeLabelIds': []
                    }
                ).execute()
                print(f"Starred email (Rule: {rule_name})")
            else:
                service.users().messages().modify(
                    userId='me',
                    id=message_id,
                    body={
                        'addLabelIds': [],
                        'removeLabelIds': ['STARRED']
                    }
                ).execute()
                print(f"Unstarred email (Rule: {rule_name})")

        elif action_lower == 'archive':
            if action_bool:
                service.users().messages().modify(
                    userId='me',
                    id=message_id,
                    body={'removeLabelIds': ['INBOX']}
                ).execute()
                print(f"Archived email (Rule: {rule_name})")
            else:
                print(f"Archive action set to False; no action taken (Rule: {rule_name})")

        elif action_lower == 'delete':
            if action_bool:
                service.users().messages().trash(userId='me', id=message_id).execute()
                print(f"Deleted email (Rule: {rule_name})")
            else:
                print(f"Delete action set to False; no action taken (Rule: {rule_name})")

        elif action_lower == 'move':
            label_id = get_or_create_label(service, action_value)
            if label_id:
                # Check if the rule targets the tag "Algorithm Reviewed [Pending]"
                pending_targeted = any(
                    cond['type'].lower() == 'tag' and cond['value'].strip().lower() == 'algorithm reviewed [pending]'
                    for cond in rule.get('conditions', [])
                )
                if pending_targeted:
                    # Retrieve the pending and reviewed tag IDs.
                    pending_label_id = get_or_create_label(service, 'Algorithm Reviewed [Pending]')
                    reviewed_label_id = get_or_create_label(service, 'Algorithm Reviewed')
                    service.users().messages().modify(
                        userId='me',
                        id=message_id,
                        body={
                            'addLabelIds': [label_id, reviewed_label_id],
                            'removeLabelIds': ['INBOX', pending_label_id]
                        }
                    ).execute()
                    print(f"Moved email to: {action_value} and replaced 'Algorithm Reviewed [Pending]' with 'Algorithm Reviewed' (Rule: {rule_name})")
                else:
                    service.users().messages().modify(
                        userId='me',
                        id=message_id,
                        body={
                            'addLabelIds': [label_id],
                            'removeLabelIds': ['INBOX']
                        }
                    ).execute()
                    print(f"Moved email to: {action_value} (Rule: {rule_name})")
                    
        else:
            if action_lower == 'label':
                label_id = get_or_create_label(service, action_value)
                if label_id:
                    service.users().messages().modify(
                        userId='me',
                        id=message_id,
                        body={'addLabelIds': [label_id]}
                    ).execute()
                    print(f"Applied label: {action_value} (Rule: {rule_name})")
    except Exception as e:
        print(f"Error performing action {action} for rule {rule_name}: {str(e)}")


def extract_attachments(msg):
    """Extract a list of attachment filenames from the email message payload."""
    attachments = []
    payload = msg.get('payload', {})
    if 'parts' in payload:
        for part in payload['parts']:
            filename = part.get('filename')
            if filename:
                attachments.append(filename)
    return attachments


def clear_failed_to_review():
    """
    Removes 'Algorithm Failed to Review' label from all emails in the inbox.
    (In case you decide to re-run them or simply want to unmark them.)
    """
    service = get_gmail_service()
    failed_label = get_or_create_label(service, 'Algorithm Failed to Review')
    if not failed_label:
        print("No 'Algorithm Failed to Review' label found or created.")
        return

    # Query all inbox emails that have the label "Algorithm Failed to Review"
    query = 'in:inbox label:"Algorithm Failed to Review"'
    results = service.users().messages().list(userId='me', q=query).execute()
    messages = results.get('messages', [])

    if not messages:
        print("No emails with 'Algorithm Failed to Review' found in inbox.")
        return

    for msg_info in messages:
        msg_id = msg_info['id']
        service.users().messages().modify(
            userId='me',
            id=msg_id,
            body={'removeLabelIds': [failed_label]}
        ).execute()

    print("All 'Algorithm Failed to Review' labels removed from inbox emails.")


def process_emails_main():
    """
    MAIN Routine:
      - Queries only emails that are unreviewed (i.e. not labeled as "Algorithm Reviewed",
        "Algorithm Reviewing [Clearance]", or "Algorithm Failed to Review").
      - Evaluates main rules; if a rule matches, the email is processed and labeled
        "Algorithm Reviewed". If no rule matches, the email is labeled "Algorithm Reviewing [Clearance]".
    """
    print("\n=== Starting MAIN Processing ===")
    service = get_gmail_service()
    main_rules = load_rules_from_excel('resources/rules.xlsx', sheet_name="MainSheet")
    print(f"Loaded {len(main_rules)} main rules from Excel")
    
    reviewed_label_id = get_or_create_label(service, 'Algorithm Reviewed')
    reviewing_clearance_id = get_or_create_label(service, 'Algorithm Reviewing [Clearance]')
    pending_label_id = get_or_create_label(service, 'Algorithm Reviewed [Pending]')
    
    query = (
        'in:inbox category:primary '
        '-label:"Algorithm Reviewed" '
        '-label:"Algorithm Failed to Review" ' 
        '-label:"Keep in Email"'
    )
    results = service.users().messages().list(userId='me', q=query).execute()
    messages = results.get('messages', [])

    if not messages:
        print("No new emails for MAIN to process")
        return 0
    
    total_messages = len(messages)
    print(f"Found {total_messages} emails to check (MAIN)")
    
    rules_applied_count = 0

    for index, msg_info in enumerate(messages, 1):
        msg_id = msg_info['id']
        print(f"\nMAIN: Processing email {index}/{total_messages} (ID: {msg_id})")
        msg = service.users().messages().get(userId='me', id=msg_id, format='full').execute()


        headers = msg['payload'].get('headers', [])
        subject = next((h['value'] for h in headers if h['name'] == 'Subject'), 'No Subject')
        sender = next((h['value'] for h in headers if h['name'] == 'From'), 'Unknown Sender')
        date_str = next((h['value'] for h in headers if h['name'] == 'Date'), None)
        email_date = parsedate_to_datetime(date_str) if date_str else datetime.now()
        full_body = get_full_body_text(msg.get('payload', {}))
        attachments = extract_attachments(msg)
        msg_labels = msg.get('labelIds', [])
        is_important = 'IMPORTANT' in msg_labels
        replies_count = 0
        thread_id = msg.get('threadId')
        if thread_id:
            thread = service.users().threads().get(userId='me', id=thread_id).execute()
            replies_count = len(thread.get('messages', [])) - 1
        else:
            replies_count = 0

        print("\n--- Email Debug Info ---")
        print(f"Subject: {subject}")
        print(f"Sender: {sender}")
        # print(f"Body excerpt: {full_body[:100]}...")
        # print(f"Labels: {msg_labels}")
        # print(f"Is Important?: {is_important}")
        # print(f"Attachments: {attachments}")
        # print(f"Replies Count: {replies_count}")

        pending_exists = pending_label_id in msg_labels
        if pending_exists:
            print("Email has 'Algorithm Reviewed [Pending]' label; only specific rules will be applied")
        
        rule_matched = False
        for rule in main_rules:
            # if rule.get('rule_type') == 'paused':
            #     continue
            if pending_exists:
                has_pending_condition = any(
                    c['type'].lower() == 'tag' and c['value'].lower() == 'algorithm reviewed [pending]'
                    for c in rule['conditions']
                )
                if not has_pending_condition:
                    continue

            
            if evaluate_conditions(rule['conditions'], subject, sender, full_body,
                                   email_date, msg_labels, service, is_important,
                                   attachments, replies_count):
                print(f"✓ MAIN Rule '{rule['name']}' matched!")
                handle_rule_action(service, msg_id, rule)
                rule_matched = True
                rules_applied_count += 1


                updated_msg = service.users().messages().get(userId='me', id=msg_id, format='metadata').execute()
                updated_labels = updated_msg.get('labelIds', [])
                if reviewed_label_id not in updated_labels and pending_label_id not in updated_labels:
                    service.users().messages().modify(
                        userId='me',
                        id=msg_id,
                        body={'addLabelIds': [reviewed_label_id]}
                    ).execute()
                    print("Added 'Algorithm Reviewed' label (Main process).")
                else:
                    print("Skipped adding 'Algorithm Reviewed' label because one already exists.")
                break
            # else:
            #     print(f"✗ Rule '{rule['name']}' did not match")

        if not rule_matched:
            print("No MAIN rule matched; labeling 'Algorithm Reviewing [Clearance]'")
            service.users().messages().modify(
                userId='me',
                id=msg_id,
                body={'addLabelIds': [reviewing_clearance_id]}
            ).execute()

            
    print(f"\n=== MAIN Processing Complete ===")
    print(f"Applied rules to {rules_applied_count} / {total_messages} emails (Main process)")
    return rules_applied_count


def process_emails_clearance():
    """
    CLEARANCE Routine:
      - Processes only emails labeled "Algorithm Reviewing [Clearance]".
      - For each email, runs clearance rules.
      - If a rule matches, handles the action, re-fetches metadata, and if the email still
        has "Algorithm Reviewing [Clearance]" (and lacks "Algorithm Reviewed"), removes the former
        and adds "Algorithm Reviewed".
      - If no clearance rule matches, removes "Algorithm Reviewing [Clearance]" and adds "Algorithm Failed to Review".
    """
    print("\n=== Starting CLEARANCE Processing ===")
    service = get_gmail_service()
    clearance_rules = load_rules_from_excel('resources/rules.xlsx', sheet_name="ClearanceRules")
    print(f"Loaded {len(clearance_rules)} clearance rules from Excel")
    
    reviewed_label_id = get_or_create_label(service, 'Algorithm Reviewed')
    reviewing_clearance_id = get_or_create_label(service, 'Algorithm Reviewing [Clearance]')
    failed_review_id = get_or_create_label(service, 'Algorithm Failed to Review')
    pending_label_id = get_or_create_label(service, 'Algorithm Reviewed [Pending]')
    
    results = service.users().messages().list(
        userId='me',
        labelIds=[reviewing_clearance_id],  # Use label ID here
        q='in:inbox',  # Optional: Restrict to inbox
        includeSpamTrash=False
    ).execute()
    messages = results.get('messages', [])

    if not messages:
        print("No emails for CLEARANCE to process")
        return 0
    
    total_messages = len(messages)
    print(f"Found {total_messages} emails to check (CLEARANCE)")
    
    rules_applied_count = 0
    
    for index, msg_info in enumerate(messages, 1):
        msg_id = msg_info['id']
        print(f"\nCLEARANCE: Processing email {index}/{total_messages} (ID: {msg_id})")
        msg = service.users().messages().get(userId='me', id=msg_id, format='full').execute()


        headers = msg['payload'].get('headers', [])
        subject = next((h['value'] for h in headers if h['name'] == 'Subject'), 'No Subject')
        sender = next((h['value'] for h in headers if h['name'] == 'From'), 'Unknown Sender')
        date_str = next((h['value'] for h in headers if h['name'] == 'Date'), None)
        email_date = parsedate_to_datetime(date_str) if date_str else datetime.now()
        full_body = get_full_body_text(msg.get('payload', {}))
        attachments = extract_attachments(msg)
        msg_labels = msg.get('labelIds', [])
        is_important = 'IMPORTANT' in msg_labels
        replies_count = 0
        thread_id = msg.get('threadId')
        if thread_id:
            thread = service.users().threads().get(userId='me', id=thread_id).execute()
            replies_count = len(thread.get('messages', [])) - 1
        else:
            replies_count = 0

        print("\n--- Email Debug Info ---")
        print(f"Subject: {subject}")
        print(f"Sender: {sender}")
        # print(f"Body excerpt: {full_body[:100]}...")
        # print(f"Labels: {msg_labels}")
        # print(f"Is Important?: {is_important}")
        # print(f"Attachments: {attachments}")
        # print(f"Replies Count: {replies_count}")

        pending_exists = pending_label_id in msg_labels
        if pending_exists:
            print("Email has 'Algorithm Reviewed [Pending]' label; only specific rules will be applied")

        rule_matched = False
        for rule in clearance_rules:
            # if rule.get('rule_type') == 'paused':
            #     continue
            if pending_exists:
                has_pending_condition = any(
                    c['type'].lower() == 'tag' and c['value'].lower() == 'algorithm reviewed [pending]'
                    for c in rule['conditions']
                )
                if not has_pending_condition:
                    continue


            if evaluate_conditions(rule['conditions'], subject, sender, full_body,
                                   email_date, msg_labels, service, is_important,
                                   attachments, replies_count):
                print(f"✓ CLEARANCE Rule '{rule['name']}' matched!")
                handle_rule_action(service, msg_id, rule)
                rule_matched = True
                rules_applied_count += 1

                
                updated_msg = service.users().messages().get(userId='me', id=msg_id, format='metadata').execute()
                updated_labels = updated_msg.get('labelIds', [])
                if reviewing_clearance_id in updated_labels and reviewed_label_id not in updated_labels:
                    service.users().messages().modify(
                        userId='me',
                        id=msg_id,
                        body={'removeLabelIds': [reviewing_clearance_id],
                              'addLabelIds': [reviewed_label_id]}
                    ).execute()
                    print("Replaced 'Algorithm Reviewing [Clearance]' with 'Algorithm Reviewed' (Clearance process).")
                else:
                    print("Skipped label update in Clearance process; label already updated.")
                break
        if not rule_matched:
            print("No CLEARANCE rule matched; labeling 'Algorithm Failed to Review'")
            service.users().messages().modify(
                userId='me',
                id=msg_id,
                body={'removeLabelIds': [reviewing_clearance_id],
                      'addLabelIds': [failed_review_id]}
            ).execute()
    print(f"\n=== CLEARANCE Processing Complete ===")
    print(f"Applied clearance rules to {rules_applied_count} / {total_messages} emails (Clearance process)")
    return rules_applied_count


def run_cleaning():
    """
    Runs the cleaning routine:
      - Repeatedly run MAIN processing.
      - When MAIN processes 0 emails, run CLEARANCE processing.
      - If CLEARANCE also processes 0, then exit.
    """
    print("=== CLEANING ROUTINE START ===")
    while True:
        main_processed = process_emails_main()
        clearance_processed = process_emails_clearance()
        if clearance_processed == 0:
            print("CLEANING DONE: No more emails to process in MAIN or CLEARANCE.")
            break
        else:
            print("Cleared some emails in CLEARANCE. Continuing cleaning routine...\n")
    print("=== CLEANING ROUTINE END ===")
    return


def clear_failed_to_review():
    """
    Removes 'Algorithm Failed to Review' label from all emails in the inbox.
    """
    service = get_gmail_service()
    failed_label_id = get_or_create_label(service, 'Algorithm Failed to Review')
    if not failed_label_id:
        print("No 'Algorithm Failed to Review' label found or created.")
        return
    query = 'in:inbox label:"Algorithm Failed to Review"'
    results = service.users().messages().list(userId='me', q=query).execute()
    messages = results.get('messages', [])
    if not messages:
        print("No emails with 'Algorithm Failed to Review' found in inbox.")
        return
    for msg_info in messages:
        msg_id = msg_info['id']
        service.users().messages().modify(
            userId='me',
            id=msg_id,
            body={'removeLabelIds': [failed_label_id]}
        ).execute()
    print("All 'Algorithm Failed to Review' labels removed from inbox emails.")


def notify(title, message):
    subprocess.run([
        'terminal-notifier',
        '-title', title,
        '-message', message,
        '-sound', 'Ping'
    ])

def main():
    """
    Decide which routine to run based on command-line argument:
      python main.py main       => run process_emails_main()
      python main.py cleaning   => run run_cleaning()
    Default if no argument => run_cleaning()
    """
    if len(sys.argv) > 1:
        mode = sys.argv[1].lower()
        if mode == "main":
            process_emails_main()
        elif mode == "cleaning":
            run_cleaning()
        else:
            print(f"Unknown mode '{mode}'. Running main routine by default.")
            process_emails_main()
    else:
        process_emails_main()
    
    notify("Email Cleaning", "Email cleaning process done!")
    print("All routines finished.")

if __name__ == "__main__":
    main()